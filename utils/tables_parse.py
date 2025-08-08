import csv

from aiogram import Bot
from aiogram.types import Message
import openpyxl
from typing import List
from pathlib import Path


def load_usernames_from_csv(
    file_path: str,
) -> List[str]:
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"CSV файл не найден: {file_path}")

    usernames = []

    with open(file_path, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        for row in reader:
            if len(row) > 0:
                username = row[0].strip()
                if username:
                    if not username.startswith('@'):
                        username = '@' + username
                    usernames.append(username)
    return usernames


def load_usernames_from_excel(
    file_path: str,
) -> List[str]:
    file_path = Path(file_path)
    if not file_path.exists():
        raise FileNotFoundError(f"Excel файл не найден: {file_path}")

    workbook = openpyxl.load_workbook(filename=file_path, read_only=True)
    sheet = workbook.active

    usernames = []
    for row in sheet.iter_rows(values_only=True):
        if len(row) > 0:
            cell_value = row[0]
            if cell_value is not None:
                username = str(cell_value).strip()
                if username:
                    if not username.startswith('@'):
                        username = '@' + username
                    usernames.append(username)

    workbook.close()
    return usernames


async def load_usernames(
    bot: Bot,
    message: Message,
) -> List[str]:
    file = await bot.get_file(message.document.file_id)
    file_path = await bot.download_file(file.file_path)
    path = Path(file_path.name)
    if not path.exists():
        raise FileNotFoundError(f"Файл не найден: {path}")

    if path.suffix.lower() == '.csv':
        return load_usernames_from_csv(file_path.name)

    elif path.suffix.lower() in ['.xlsx']:
        return load_usernames_from_excel(file_path.name)
    else:
        raise ValueError(f"Неподдерживаемый формат: {path.suffix}. Поддерживаются: .csv, .xlsx")


def get_table(tables: list[str], name: str) -> str:
    """
        Возвращает путь к файлу таблицы
    """
    wb = openpyxl.Workbook()
    sheet = wb.active

    for column in range(0, len(tables)):
        c = sheet.cell(row=column, column=0)
        c.value = tables[column]
    wb.save(f'{name}.xlsx')
    return f'{name}.xlsx'
